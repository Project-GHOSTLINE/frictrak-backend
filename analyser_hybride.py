#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
SAR Analysis - Systeme Hybride v2.0
===================================

Systeme d'analyse automatique Solution Argent Rapide
avec detection hybride (Liste OPC + Regles intelligentes)

Version: 2.0 Hybride
Date: 2025-11-28

FONCTIONNALITES:
- Detection via liste officielle OPC (414 preteurs) = 95%+ precision
- Detection via regles intelligentes pour cas non couverts
- Elimination automatique des faux positifs (assurances, casinos, syndics)
- Scoring de confiance 0-100%
- Rapports detailles avec source de detection
"""

import os
import json
import glob
import re
import sys
from datetime import datetime, timedelta
from collections import defaultdict
from openpyxl import load_workbook

# =============================================================================
# IMPORT DES LISTES OFFICIELLES OPC
# =============================================================================

from listes_completes_v2 import (
    PRETEURS_TOUS,
    PRETEURS_OFFICIELS_OPC,
    PRETEURS_COMPLEMENTAIRES,
    ASSURANCES_TOUTES,
    SYNDICS_TOUS,
    CASINOS_JEUX,
    COMMERCES_SERVICES,
    categoriser_transaction,
    est_preteur,
    est_a_exclure,
    est_assurance,
    est_syndic,
    est_casino
)

# =============================================================================
# CONFIGURATION
# =============================================================================

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
JSON_INPUT_DIR = os.path.join(SCRIPT_DIR, "json-input")
RAPPORTS_OUTPUT_DIR = os.path.join(SCRIPT_DIR, "rapports-output")
EXCEL_FILE = os.path.join(SCRIPT_DIR, "tableau_paiements.xlsx")

# =============================================================================
# CONSTANTES POUR LE SCORING
# =============================================================================

# Mots-cles suspects indiquant un preteur (+30 points)
KEYWORDS_PRETEUR = [
    # Francais
    "PRET", "PRÊT", "EMPRUNT", "AVANCE", "CREDIT", "CRÉDIT",
    "FINANCE", "FINANCIERE", "FINANCIÈRE", "FINANCEMENT",
    "ARGENT", "CASH", "COMPTANT", "RAPIDE", "EXPRESS",
    # Anglais
    "LOAN", "LOANS", "LENDING", "LENDER", "ADVANCE",
    "MONEY", "FAST", "QUICK", "INSTANT", "PAYDAY",
    # Prefixes transactions
    "PMT", "PAYMENT", "PAIEMENT", "VIR", "VIREMENT", "TRANSFER"
]

# Montants ronds suspects (+15 points)
MONTANTS_SUSPECTS = [
    100, 150, 200, 250, 300, 400, 500,
    750, 1000, 1500, 2000, 2500, 3000,
    4000, 5000
]

# Prefixes bancaires suspects (+10 points)
PREFIXES_SUSPECTS = [
    "PMT", "PAIEMENT", "PAYMENT",
    "VIR", "VIREMENT", "TRANSFER",
    "RETRAIT", "WITHDRAWAL", "DEBIT"
]

# Mots indiquant NON-preteur (-40 points)
NON_PRETEUR_KEYWORDS = [
    # Salaire
    "SALAIRE", "SALARY", "PAIE", "PAYROLL", "WAGES",
    "REMUNERATION", "RÉMUNÉRATION",
    # Remboursements fiscaux
    "REMBOURSEMENT IMPOT", "TAX REFUND", "IMPOT", "TAX RETURN",
    "REVENU QUEBEC", "REVENU CANADA",
    # Virements personnels / Interac (PAS des preteurs!)
    "VIREMENT PERSONNEL", "TRANSFER FROM", "DEPOT", "DEPOSIT",
    "TRANSFER PERSONAL", "PERSONNEL",
    "VIR INTERAC ENVOYE", "VIR INTERAC EFFECTUE", "VIREMENT INTERAC",
    "VIR INTERAC",  # Catch-all pour tous les virements Interac
    # Autres revenus
    "DIVIDEND", "DIVIDENDE", "INTEREST", "INTERET", "INTÉRÊT",
    "PENSION", "RETIREMENT",
    # Ventes
    "VENTE", "SALE", "SOLD", "REFUND"
]

# Descriptions trop generiques
GENERIC_DESCRIPTIONS = [
    "TRANSFER", "VIREMENT", "PAYMENT", "PAIEMENT",
    "DEBIT", "CREDIT", "RETRAIT", "DEPOT"
]

# Mots-cles paie (pour detection revenu)
PAYROLL_KEYWORDS = [
    "PAIE", "PAYROLL", "SALARY", "SALAIRE", "PAY", "WAGE",
    "DIRECT DEPOSIT", "DEPOT DIRECT", "REMUNERATION"
]

# Mots-cles SAR
SAR_KEYWORDS = ["SAR", "SOLUTION ARGENT", "ARGENT RAPIDE"]

# =============================================================================
# EXCLUSIONS: Casinos, gambling et services non-prêteurs
# =============================================================================

EXCLUSIONS_GAMBLING = [
    'GIGADAT',
    'LOTO-QUEBEC',
    'LOTOQUEBEC',
    'ESPACEJEUX',
    'CASINO',
    'MISE-O-JEU',
    'POKER',
    'SLOTS',
    'BINGO',
    'PLAYNOW',
    'BET',
    'GAMING'
]

EXCLUSIONS_SERVICES_PAIEMENT = [
    'LOONIO',
    'PAYPER',
    'KOHO',
    'WEALTHSIMPLE',
    'QUESTRADE',
    'TANGERINE',
    'EQ BANK'
]

EXCLUSIONS_AUTRES = [
    'IGA',
    'METRO',
    'WALMART',
    'COSTCO',
    'DOLLARAMA',
    'CANADIAN TIRE',
    'SAAQ',
    'HYDRO',
    'VIDEOTRON',
    'BELL',
    'ROGERS',
    'TIM HORTONS',
    'MCDONALD',
    'JEAN COUTU'
]

# NOUVEAU: Patterns de transactions à exclure (pas des prêteurs)
EXCLUSIONS_PATTERNS = [
    'BILL PAYMENT',
    'TRANSFERSBILL',
    'TRANSFERSTRANSFER',
    'TRANSFERSINTERAC E-TRANSFER TO',
    'E-TRANSFER TO',
    'INTERAC E-TRANSFER TO',
    'WITHDRAWAL AT THE COUNTER',
    'WITHDRAWAL ON ATM',
    'CHEQUES/CASHWITHDRAWAL',
    'TRANSFERSWITHDRAWAL',
    'DEPOSIT ON ATM',
    'MOBILE DEPOSIT',
    'MASTERCARD',
    'VISA',
    'CARTE DE CRÉDIT',
    'CREDIT CARD',
    'DESJARDINS ODYSSÉE',
    'BMO',
    'RBC',
    'CIBC',
    'SCOTIA',
    'A30 EXPRESS',
    'A25 EXPRESS',
    'RQ PAYMENT',
    'MUNICIPAL',
    'SCHOOL TAXES'
]

def est_exclus(nom, categorie=''):
    """
    Vérifie si une transaction doit être exclue de la détection de prêteurs.

    Args:
        nom: Nom du merchant/destinataire
        categorie: Catégorie Inverite (ex: "entertainment/gambling")

    Returns:
        True si doit être exclu, False sinon
    """
    nom_upper = nom.upper()

    # 1. PRIORITÉ: Vérifier catégorie Inverite
    if categorie:
        categorie_lower = categorie.lower()

        # Exclure gambling/casino
        if 'gambling' in categorie_lower or 'casino' in categorie_lower:
            print(f"   ⚠️ EXCLU (gambling): {nom}")
            return True

        # Exclure groceries, gas, utilities
        if any(x in categorie_lower for x in ['groceries', 'gas', 'fuel', 'utilities', 'bills', 'insurance/car', 'insurance/life']):
            return True

        # Exclure ATM (mais PAS les transfers - peuvent être des paiements prêteurs)
        if 'atm' in categorie_lower:
            print(f"   ⚠️ EXCLU (ATM): {nom}")
            return True

        # Pour les transfers, vérifier si c'est un prêteur connu AVANT d'exclure
        if 'transfer' in categorie_lower:
            # Ne PAS exclure si le nom contient un prêteur connu
            from listes_completes_v2 import est_preteur
            is_lender, _ = est_preteur(nom)
            if not is_lender:
                # Vérifier aussi les mots-clés prêteurs
                nom_upper = nom.upper()
                lender_keywords = ['CREDIT', 'LOAN', 'PRET', 'PRÊT', 'FINANCE', 'MONEY', 'CASH', 'SECOURS']
                if not any(kw in nom_upper for kw in lender_keywords):
                    print(f"   ⚠️ EXCLU (transfer non-prêteur): {nom}")
                    return True

    # 2. CRITIQUE: Exclure patterns de transactions bancaires
    for pattern in EXCLUSIONS_PATTERNS:
        if pattern in nom_upper:
            print(f"   ⚠️ EXCLU (pattern bancaire): {nom}")
            return True

    # 3. Vérifier listes d'exclusions hardcodées
    toutes_exclusions = EXCLUSIONS_GAMBLING + EXCLUSIONS_SERVICES_PAIEMENT + EXCLUSIONS_AUTRES + EXCLUSIONS_PATTERNS

    for exclusion in toutes_exclusions:
        if exclusion in nom_upper:
            return True

    # 4. Exclure noms trop courts (< 4 caractères)
    if len(nom.strip()) < 4:
        return True

    return False


# =============================================================================
# FONCTIONS UTILITAIRES
# =============================================================================

def count_similar_transactions(description, transaction_history, threshold=0.8):
    """
    Compte les transactions similaires dans l'historique

    Args:
        description: Description de la transaction actuelle
        transaction_history: Liste des transactions passees
        threshold: Similarite minimale (80% par defaut)

    Returns:
        int: Nombre de transactions similaires
    """
    if not transaction_history:
        return 0

    similar_count = 0
    desc_words = set(description.upper().split())

    if len(desc_words) == 0:
        return 0

    for past_trans in transaction_history:
        # Supporter differents formats de transaction
        if isinstance(past_trans, dict):
            past_desc = past_trans.get('description', past_trans.get('Description', ''))
        else:
            past_desc = str(past_trans)

        past_words = set(past_desc.upper().split())

        if len(past_words) == 0:
            continue

        # Calcul similarite Jaccard
        intersection = len(desc_words & past_words)
        union = len(desc_words | past_words)

        if union > 0:
            similarity = intersection / union
            if similarity >= threshold:
                similar_count += 1

    return similar_count


def normalize_description(description):
    """Normalise une description pour comparaison"""
    if not description:
        return ""
    # Majuscules, retirer caracteres speciaux
    normalized = description.upper()
    normalized = re.sub(r'[^\w\s]', ' ', normalized)
    normalized = re.sub(r'\s+', ' ', normalized).strip()
    return normalized


# =============================================================================
# FONCTION PRINCIPALE: SCORING HYBRIDE
# =============================================================================

def calculate_lender_score_hybrid(description, amount, transaction_history=None):
    """
    Calcule le score d'un preteur avec approche hybride

    SYSTEME DE SCORING:
    ===================
    Score final = Base (Liste) + Regles positives - Penalites
    Normalise entre 0-100

    Args:
        description: Description de la transaction
        amount: Montant de la transaction (valeur absolue)
        transaction_history: Liste des transactions passees (optionnel)

    Returns:
        dict: {
            'score': int (0-100),
            'confidence': str ('TRES ELEVEE', 'ELEVEE', 'MOYENNE', 'FAIBLE', 'TRES FAIBLE'),
            'source': str ('LISTE_OPC', 'LISTE_COMPLEMENTAIRE', 'REGLES', 'EXCLUSION', 'AUCUNE'),
            'reasons': list[str],
            'preteur_nom': str (si trouve),
            'action': str (recommandation)
        }
    """

    if transaction_history is None:
        transaction_history = []

    # Valeur absolue du montant
    amount = abs(amount) if amount else 0
    desc_upper = normalize_description(description)

    # =========================================================================
    # ETAPE 1: VERIFICATION LISTE OFFICIELLE (PRIORITE ABSOLUE)
    # =========================================================================

    result = categoriser_transaction(description)

    # Preteur officiel confirme
    if result['type'] == 'preteur':
        # Determiner si OPC ou complementaire
        source = 'LISTE_OPC'
        if result['nom'] in PRETEURS_COMPLEMENTAIRES:
            source = 'LISTE_COMPLEMENTAIRE'

        return {
            'score': 95 if source == 'LISTE_OPC' else 90,
            'confidence': 'TRES ELEVEE',
            'source': source,
            'reasons': [f"Preteur officiel identifie: {result['nom']}"],
            'preteur_nom': result['nom'],
            'action': 'CONFIRMER - Preteur officiel du gouvernement du Quebec'
        }

    # Exclusion automatique (assurance/casino/syndic/commerce)
    if result['doit_exclure']:
        return {
            'score': 0,
            'confidence': 'N/A',
            'source': 'EXCLUSION',
            'reasons': [f"Exclusion automatique: {result['type'].upper()} ({result['nom']})"],
            'preteur_nom': '',
            'action': f"EXCLURE - {result['type'].capitalize()} identifie"
        }

    # =========================================================================
    # ETAPE 2: VERIFIER NON-PRETEURS EN PRIORITE
    # =========================================================================

    # Si clairement un non-preteur, retourner 0 immediatement
    for non_kw in NON_PRETEUR_KEYWORDS:
        if non_kw in desc_upper:
            return {
                'score': 0,
                'confidence': 'TRES FAIBLE',
                'source': 'AUCUNE',
                'reasons': [f"Non-preteur detecte: {non_kw}"],
                'preteur_nom': '',
                'action': 'NON-PRETEUR - Ignorer'
            }

    # =========================================================================
    # ETAPE 3: APPLIQUER REGLES INTELLIGENTES
    # =========================================================================

    score = 0
    reasons = []

    # -------------------------------------------------------------------------
    # REGLE A: Mots-cles suspects (+30 points)
    # -------------------------------------------------------------------------
    keywords_found = []
    for keyword in KEYWORDS_PRETEUR:
        if keyword in desc_upper:
            keywords_found.append(keyword)

    if keywords_found:
        score += 30
        reasons.append(f"Mots-cles suspects: {', '.join(keywords_found[:3])}")

    # -------------------------------------------------------------------------
    # REGLE B: Montants ronds suspects (+15 points)
    # -------------------------------------------------------------------------
    if amount in MONTANTS_SUSPECTS:
        score += 15
        reasons.append(f"Montant rond suspect: ${amount:,.0f}")

    # -------------------------------------------------------------------------
    # REGLE C: Transactions repetitives (+10 a +35 points)
    # -------------------------------------------------------------------------
    similar_count = count_similar_transactions(description, transaction_history)

    if similar_count >= 5:
        score += 35
        reasons.append(f"Transactions tres repetitives ({similar_count} similaires)")
    elif similar_count >= 3:
        score += 25
        reasons.append(f"Transactions repetitives ({similar_count} similaires)")
    elif similar_count >= 2:
        score += 10
        reasons.append(f"Quelques repetitions ({similar_count} similaires)")

    # -------------------------------------------------------------------------
    # REGLE D: Prefixes bancaires (+10 points)
    # -------------------------------------------------------------------------
    for prefix in PREFIXES_SUSPECTS:
        if desc_upper.startswith(prefix):
            score += 10
            reasons.append(f"Prefixe bancaire suspect: {prefix}")
            break

    # -------------------------------------------------------------------------
    # REGLE E: Montants typiques payday loans (+10 points)
    # -------------------------------------------------------------------------
    if 50 <= amount <= 5000:
        score += 10
        reasons.append("Montant typique payday loan (50$-5000$)")

    # -------------------------------------------------------------------------
    # REGLE F: Combinaisons suspectes (+20 a +25 points)
    # -------------------------------------------------------------------------

    # Combinaison 1: "CREDIT" + montant < 2000$
    if "CREDIT" in desc_upper and amount < 2000 and amount > 0:
        score += 20
        reasons.append("Combinaison: CREDIT + petit montant")

    # Combinaison 2: "FINANCE" + montant rond
    if "FINANCE" in desc_upper and amount in MONTANTS_SUSPECTS:
        score += 20
        reasons.append("Combinaison: FINANCE + montant rond")

    # Combinaison 3: "LOAN" + repetitions
    if "LOAN" in desc_upper and similar_count >= 2:
        score += 20
        reasons.append("Combinaison: LOAN + repetitions")

    # Combinaison 4: Triple combinaison (prefixe + keyword + montant)
    has_prefix = any(desc_upper.startswith(p) for p in ["PMT", "VIR"])
    has_keyword = any(kw in desc_upper for kw in ["CREDIT", "LOAN", "FINANCE", "PRET"])
    has_amount = 50 <= amount <= 5000

    if has_prefix and has_keyword and has_amount:
        score += 25
        reasons.append("Triple combinaison suspecte (prefixe + mot-cle + montant)")

    # =========================================================================
    # ETAPE 4: APPLIQUER PENALITES
    # =========================================================================

    # -------------------------------------------------------------------------
    # PENALITE A: Montants tres eleves (-20 a -40 points)
    # -------------------------------------------------------------------------
    if amount > 20000:
        score -= 40
        reasons.append("Montant extreme (>20K$)")
    elif amount > 10000:
        score -= 20
        reasons.append("Montant tres eleve (>10K$)")

    # -------------------------------------------------------------------------
    # PENALITE B: Descriptions generiques (-15 a -20 points)
    # -------------------------------------------------------------------------
    if len(description.strip()) < 10:
        score -= 15
        reasons.append("Description trop courte")

    if description.upper().strip() in GENERIC_DESCRIPTIONS:
        score -= 20
        reasons.append("Description trop generique")

    # -------------------------------------------------------------------------
    # PENALITE C: Frequence excessive (-10 points)
    # -------------------------------------------------------------------------
    if similar_count > 20:
        score -= 10
        reasons.append("Frequence excessive (>20 transactions)")

    # =========================================================================
    # ETAPE 5: NORMALISER ET DETERMINER CONFIANCE
    # =========================================================================

    final_score = max(0, min(100, score))

    # Definir niveau de confiance
    if final_score >= 80:
        confidence = "TRES ELEVEE"
        action = "CONFIRMER - Tres probable preteur"
    elif final_score >= 60:
        confidence = "ELEVEE"
        action = "PROBABLE - Verifier manuellement"
    elif final_score >= 40:
        confidence = "MOYENNE"
        action = "POSSIBLE - Investigation requise"
    elif final_score >= 20:
        confidence = "FAIBLE"
        action = "IMPROBABLE - Peut ignorer"
    else:
        confidence = "TRES FAIBLE"
        action = "NON-PRETEUR - Ignorer"

    # Determiner la source
    if final_score > 0 and len(reasons) > 0:
        source = 'REGLES'
    else:
        source = 'AUCUNE'

    return {
        'score': final_score,
        'confidence': confidence,
        'source': source,
        'reasons': reasons,
        'preteur_nom': '',  # Pas dans liste officielle
        'action': action
    }


# =============================================================================
# DETECTION DES PRETEURS DANS LES TRANSACTIONS
# =============================================================================

def detect_lenders_hybrid(transactions):
    """
    Detecte tous les preteurs dans une liste de transactions
    avec approche hybride

    Args:
        transactions: Liste de dicts avec keys: date, description, amount, type

    Returns:
        dict: {
            'preteurs_confirmes': [...],  # Liste OPC
            'preteurs_probables': [...],  # Score >= 60
            'preteurs_possibles': [...],  # Score 40-59
            'exclusions': [...],          # Assurances, casinos, etc.
            'statistiques': {...}
        }
    """

    preteurs_confirmes = {}   # Cle = nom preteur
    preteurs_probables = {}
    preteurs_possibles = {}
    exclusions = {}

    # Construire l'historique pour detection repetitions
    transaction_history = transactions.copy()

    for trans in transactions:
        description = trans.get('description', trans.get('Description', ''))
        amount = trans.get('amount', trans.get('Montant', 0))
        date = trans.get('date', trans.get('Date', ''))
        trans_type = trans.get('type', trans.get('Type', ''))

        # Seulement analyser les debits (paiements vers preteurs)
        if amount > 0:
            continue  # Credits = depots, pas des paiements preteurs

        amount = abs(amount)

        # NOUVEAU: Récupérer catégorie Inverite
        categorie = trans.get('category', '')

        # NOUVEAU: Exclure si gambling, groceries, etc.
        if est_exclus(description, categorie):
            continue

        # Calculer le score
        result = calculate_lender_score_hybrid(description, amount, transaction_history)

        score = result['score']
        source = result['source']
        preteur_nom = result['preteur_nom'] or description[:40]

        # Creer entry pour le preteur
        entry = {
            'nom': preteur_nom,
            'score': score,
            'confidence': result['confidence'],
            'source': source,
            'reasons': result['reasons'],
            'action': result['action'],
            'transactions': [],
            'total_paye': 0
        }

        # Categoriser selon le score et la source
        if source == 'EXCLUSION':
            key = preteur_nom
            if key not in exclusions:
                exclusions[key] = entry
            exclusions[key]['transactions'].append({
                'date': date,
                'description': description,
                'amount': amount
            })
            exclusions[key]['total_paye'] += amount

        elif source in ['LISTE_OPC', 'LISTE_COMPLEMENTAIRE']:
            key = preteur_nom
            if key not in preteurs_confirmes:
                preteurs_confirmes[key] = entry
            preteurs_confirmes[key]['transactions'].append({
                'date': date,
                'description': description,
                'amount': amount
            })
            preteurs_confirmes[key]['total_paye'] += amount

        elif score >= 60:
            # Grouper par description similaire
            key = normalize_description(description)[:30]
            if key not in preteurs_probables:
                preteurs_probables[key] = entry
            preteurs_probables[key]['transactions'].append({
                'date': date,
                'description': description,
                'amount': amount
            })
            preteurs_probables[key]['total_paye'] += amount
            # Mettre a jour le score max
            if score > preteurs_probables[key]['score']:
                preteurs_probables[key]['score'] = score
                preteurs_probables[key]['confidence'] = result['confidence']
                preteurs_probables[key]['reasons'] = result['reasons']

        elif score >= 40:
            key = normalize_description(description)[:30]
            if key not in preteurs_possibles:
                preteurs_possibles[key] = entry
            preteurs_possibles[key]['transactions'].append({
                'date': date,
                'description': description,
                'amount': amount
            })
            preteurs_possibles[key]['total_paye'] += amount
            if score > preteurs_possibles[key]['score']:
                preteurs_possibles[key]['score'] = score
                preteurs_possibles[key]['confidence'] = result['confidence']
                preteurs_possibles[key]['reasons'] = result['reasons']

    # Statistiques
    total_dette_confirmee = sum(p['total_paye'] for p in preteurs_confirmes.values())
    total_dette_probable = sum(p['total_paye'] for p in preteurs_probables.values())
    total_dette_possible = sum(p['total_paye'] for p in preteurs_possibles.values())

    statistiques = {
        'nb_preteurs_confirmes': len(preteurs_confirmes),
        'nb_preteurs_probables': len(preteurs_probables),
        'nb_preteurs_possibles': len(preteurs_possibles),
        'nb_exclusions': len(exclusions),
        'total_transactions_analysees': len(transactions),
        'dette_confirmee': total_dette_confirmee,
        'dette_probable': total_dette_probable,
        'dette_possible': total_dette_possible,
        'dette_totale_estimee': total_dette_confirmee + total_dette_probable
    }

    return {
        'preteurs_confirmes': list(preteurs_confirmes.values()),
        'preteurs_probables': list(preteurs_probables.values()),
        'preteurs_possibles': list(preteurs_possibles.values()),
        'exclusions': list(exclusions.values()),
        'statistiques': statistiques
    }


# =============================================================================
# FONCTIONS KNOCK-OUT (CRITÈRES ÉLIMINATOIRES)
# =============================================================================

def extraire_nsf_et_overdraft(transactions, inverite_data=None):
    """
    Extrait le nombre de NSF et overdrafts
    Returns:
        tuple: (nsf_30j, overdraft_90j)
    """
    # Si données Inverite disponibles, utiliser les statistics
    if inverite_data:
        try:
            # Chercher dans accounts[0].statistics.quarter_all_time
            accounts = inverite_data.get('accounts', [])
            if accounts:
                stats = accounts[0].get('statistics', {})
                quarter_stats = stats.get('quarter_all_time', {})

                # average_number_nsf est le nombre moyen de NSF par mois
                nsf_avg = float(quarter_stats.get('average_number_nsf', 0))

                # Arrondir au nombre entier le plus proche pour les 30 jours
                nsf_30j = round(nsf_avg)

                print(f"   📊 NSF depuis Inverite statistics: {nsf_30j} (moyenne: {nsf_avg})")
            else:
                nsf_30j = 0

        except Exception as e:
            print(f"   ⚠️ Erreur lecture NSF Inverite: {e}")
            nsf_30j = 0
    else:
        # Fallback: compter manuellement (pour compatibilité anciens formats)
        nsf_30j = 0
        date_limite_30j = datetime.now() - timedelta(days=30)

        for trans in transactions:
            description = trans.get('description', trans.get('Description', '')).upper()
            date = trans.get('date', trans.get('Date'))

            if isinstance(date, str):
                try:
                    date = datetime.strptime(date[:10], "%Y-%m-%d")
                except:
                    continue

            if date and date >= date_limite_30j:
                # NSF réels uniquement
                if any(kw in description for kw in ['NSF FEE', 'NSF CHARGE', 'FONDS INSUFFISANTS', 'INSUFFICIENT FUNDS']):
                    nsf_30j += 1

    # Overdraft: garder la détection manuelle
    overdraft_90j = 0
    date_limite_90j = datetime.now() - timedelta(days=90)

    for trans in transactions:
        description = trans.get('description', trans.get('Description', '')).upper()
        date = trans.get('date', trans.get('Date'))

        if isinstance(date, str):
            try:
                date = datetime.strptime(date[:10], "%Y-%m-%d")
            except:
                continue

        if date and date >= date_limite_90j:
            if any(kw in description for kw in ['OVERDRAFT', 'DÉCOUVERT', 'FRAIS DÉCOUVERT', 'OD FEE']):
                overdraft_90j += 1

    return nsf_30j, overdraft_90j


def verifier_knock_outs(data):
    """
    Vérifie les critères éliminatoires (KNOCK-OUTS)
    Retourne un dict avec le résultat de chaque vérification
    """

    revenus_mensuels = data.get('revenus_mensuels', 0)
    paiements_prets = data.get('total_paiements_prets', 0)
    nb_preteurs = data.get('nb_preteurs_confirmes', 0)
    nsf_30j = data.get('nsf_30_jours', 0)
    overdraft_90j = data.get('overdraft_90_jours', 0)

    knock_outs = []

    # KNOCK-OUT #1: Ratio D/R > 100%
    if revenus_mensuels > 0:
        ratio_dr = (paiements_prets / revenus_mensuels) * 100

        if ratio_dr > 200:
            knock_outs.append({
                'type': 'RATIO_DR_CRITIQUE',
                'valeur': ratio_dr,
                'seuil': 200,
                'message': f'Ratio D/R critique: {ratio_dr:.1f}% (seuil: 200%)'
            })
        elif ratio_dr > 100:
            knock_outs.append({
                'type': 'RATIO_DR_SEVERE',
                'valeur': ratio_dr,
                'seuil': 100,
                'message': f'Ratio D/R sévère: {ratio_dr:.1f}% (seuil: 100%)'
            })

    # KNOCK-OUT #2: Capacité résiduelle < 0
    if revenus_mensuels > 0:
        capacite_max = revenus_mensuels * 0.50
        capacite_residuelle = capacite_max - paiements_prets

        if capacite_residuelle < 0:
            knock_outs.append({
                'type': 'CAPACITE_NEGATIVE',
                'valeur': capacite_residuelle,
                'seuil': 0,
                'message': f'Capacité résiduelle négative: {capacite_residuelle:.2f}$ (dépasse 50% du revenu)'
            })

    # KNOCK-OUT #3: Trop de prêteurs
    if nb_preteurs > 5:
        knock_outs.append({
            'type': 'TROP_PRETEURS',
            'valeur': nb_preteurs,
            'seuil': 5,
            'message': f'Trop de prêteurs: {nb_preteurs} (seuil: 5)'
        })

    # KNOCK-OUT #4: NSF répétés
    if nsf_30j > 2:
        knock_outs.append({
            'type': 'NSF_REPETES',
            'valeur': nsf_30j,
            'seuil': 2,
            'message': f'Chèques sans provision répétés: {nsf_30j} en 30 jours (seuil: 2)'
        })

    # KNOCK-OUT #5: Découverts chroniques
    if overdraft_90j > 5:
        knock_outs.append({
            'type': 'OVERDRAFT_CHRONIQUE',
            'valeur': overdraft_90j,
            'seuil': 5,
            'message': f'Découverts chroniques: {overdraft_90j} en 90 jours (seuil: 5)'
        })

    return {
        'has_knock_outs': len(knock_outs) > 0,
        'knock_outs': knock_outs,
        'nb_knock_outs': len(knock_outs)
    }


def calculer_score_avance(data, knock_outs_result):
    """
    Calcule le score avancé avec pénalités progressives
    Si knock-outs détectés: score = 0
    Sinon: score calculé avec pénalités
    """

    # Si knock-outs: score = 0 automatiquement
    if knock_outs_result['has_knock_outs']:
        return {
            'score': 0,
            'niveau_risque': 'CRITIQUE',
            'raison': 'REFUS AUTOMATIQUE - Knock-outs déclenchés',
            'knock_outs': knock_outs_result['knock_outs'],
            'penalites': []
        }

    # Score de base
    score = 100
    penalites = []

    revenus = data.get('revenus_mensuels', 0)
    paiements = data.get('total_paiements_prets', 0)
    budget = data.get('budget_disponible', 0)
    nb_preteurs = data.get('nb_preteurs_confirmes', 0)
    nsf = data.get('nsf_30_jours', 0)
    overdraft = data.get('overdraft_90_jours', 0)

    # PÉNALITÉ #1: Ratio D/R (poids 35%)
    if revenus > 0:
        ratio_dr = (paiements / revenus) * 100

        if ratio_dr > 75:
            score -= 20
            penalites.append(f'Ratio D/R élevé ({ratio_dr:.1f}%): -20 pts')
        elif ratio_dr > 50:
            score -= 15
            penalites.append(f'Ratio D/R limite ({ratio_dr:.1f}%): -15 pts')
        elif ratio_dr > 30:
            score -= 10
            penalites.append(f'Ratio D/R modéré ({ratio_dr:.1f}%): -10 pts')
        elif ratio_dr > 15:
            score -= 5
            penalites.append(f'Ratio D/R acceptable ({ratio_dr:.1f}%): -5 pts')

    # PÉNALITÉ #2: Budget disponible (poids 25%)
    if budget < 0:
        score -= 20
        penalites.append(f'Budget négatif ({budget:.2f}$): -20 pts')
    elif budget < 500:
        score -= 15
        penalites.append(f'Budget très serré ({budget:.2f}$): -15 pts')
    elif budget < 1000:
        score -= 10
        penalites.append(f'Budget serré ({budget:.2f}$): -10 pts')
    elif budget < 1500:
        score -= 5
        penalites.append(f'Budget limite ({budget:.2f}$): -5 pts')

    # PÉNALITÉ #3: Nombre de prêteurs (poids 20%)
    if nb_preteurs == 5:
        score -= 20
        penalites.append(f'5 prêteurs actifs: -20 pts')
    elif nb_preteurs == 4:
        score -= 15
        penalites.append(f'4 prêteurs actifs: -15 pts')
    elif nb_preteurs == 3:
        score -= 10
        penalites.append(f'3 prêteurs actifs: -10 pts')
    elif nb_preteurs == 2:
        score -= 5
        penalites.append(f'2 prêteurs actifs: -5 pts')
    elif nb_preteurs == 1:
        score -= 2
        penalites.append(f'1 prêteur actif: -2 pts')

    # PÉNALITÉ #4: NSF (poids 10%)
    if nsf == 2:
        score -= 15
        penalites.append(f'2 NSF en 30j: -15 pts')
    elif nsf == 1:
        score -= 10
        penalites.append(f'1 NSF en 30j: -10 pts')

    # PÉNALITÉ #5: Overdraft (poids 10%)
    if overdraft == 5:
        score -= 12
        penalites.append(f'5 découverts en 90j: -12 pts')
    elif overdraft >= 3:
        score -= 8
        penalites.append(f'{overdraft} découverts en 90j: -8 pts')
    elif overdraft >= 1:
        score -= 5
        penalites.append(f'{overdraft} découvert(s) en 90j: -5 pts')

    # Score final (minimum 0)
    score = max(0, score)

    # Déterminer niveau de risque
    if score >= 80:
        niveau_risque = 'FAIBLE'
    elif score >= 60:
        niveau_risque = 'MODÉRÉ'
    elif score >= 40:
        niveau_risque = 'ÉLEVÉ'
    else:
        niveau_risque = 'TRÈS ÉLEVÉ'

    return {
        'score': score,
        'niveau_risque': niveau_risque,
        'penalites': penalites,
        'knock_outs': []
    }


# =============================================================================
# CALCUL PROBABILITE REALISTE
# =============================================================================

def calculer_probabilite_realiste(score_global, ratio_dette, nb_preteurs, montant_demande, capacite_max):
    """Calcule une probabilité d'approbation réaliste basée sur tous les facteurs de risque"""

    # Probabilité de base selon le score
    if score_global >= 80:
        chance_base = 95
    elif score_global >= 70:
        chance_base = 85
    elif score_global >= 60:
        chance_base = 75
    elif score_global >= 50:
        chance_base = 60
    elif score_global >= 40:
        chance_base = 40
    elif score_global >= 30:
        chance_base = 20
    else:
        chance_base = 5

    # Pénalités selon le ratio dette/revenu
    if ratio_dette > 0.8:
        chance_base -= 40
    elif ratio_dette > 0.6:
        chance_base -= 30
    elif ratio_dette > 0.5:
        chance_base -= 20
    elif ratio_dette > 0.4:
        chance_base -= 10

    # Pénalités selon le nombre de prêteurs
    if nb_preteurs >= 8:
        chance_base -= 30
    elif nb_preteurs >= 6:
        chance_base -= 20
    elif nb_preteurs >= 4:
        chance_base -= 10

    # Pénalité selon le montant demandé vs capacité
    ratio_montant = montant_demande / capacite_max if capacite_max > 0 else 1
    if ratio_montant > 0.8:
        chance_base -= 15
    elif ratio_montant > 0.6:
        chance_base -= 10
    elif ratio_montant > 0.4:
        chance_base -= 5

    # Garantir entre 0 et 100
    return max(0, min(100, chance_base))


# =============================================================================
# GENERATION DE RAPPORT ENRICHI
# =============================================================================

def generate_report_enrichi(client_info, transactions, detection_result, paiements, filename, inverite_data=None):
    """
    Genere un rapport enrichi avec source de detection et confiance

    Args:
        client_info: Dict avec infos client
        transactions: Liste des transactions
        detection_result: Resultat de detect_lenders_hybrid()
        paiements: Grille de paiements
        filename: Nom du fichier source
        inverite_data: Données complètes Inverite (optionnel)

    Returns:
        str: Rapport formate
    """

    stats = detection_result['statistiques']
    preteurs_confirmes = detection_result['preteurs_confirmes']
    preteurs_probables = detection_result['preteurs_probables']
    preteurs_possibles = detection_result['preteurs_possibles']
    exclusions = detection_result['exclusions']

    # Calculer revenu mensuel et capacite
    monthly_income = client_info.get('revenu_mensuel', 0)
    if monthly_income == 0:
        monthly_income = calculate_monthly_income_from_transactions(transactions)

    capacite_mensuelle = monthly_income * 0.50
    capacite_hebdo = capacite_mensuelle / 4.33

    # Construction du rapport
    report = []

    # En-tete
    report.append("")
    report.append("=" * 70)
    report.append("  RAPPORT D'ANALYSE SAR - SYSTEME HYBRIDE V2.0")
    report.append("  Solution Argent Rapide")
    report.append("=" * 70)
    report.append("")

    # Date et source
    report.append(f"Date du rapport: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    report.append(f"Fichier source: {filename}")
    report.append("")

    # Resume executif
    report.append("-" * 70)
    report.append("  RESUME EXECUTIF")
    report.append("-" * 70)
    report.append("")
    report.append(f"Transactions analysees: {stats['total_transactions_analysees']}")
    report.append(f"Preteurs confirmes (Liste OPC): {stats['nb_preteurs_confirmes']}")
    report.append(f"Preteurs probables (Regles): {stats['nb_preteurs_probables']}")
    report.append(f"Preteurs possibles (A verifier): {stats['nb_preteurs_possibles']}")
    report.append(f"Exclusions (Faux positifs elimines): {stats['nb_exclusions']}")
    report.append("")
    report.append(f"Dette confirmee: ${stats['dette_confirmee']:,.2f}")
    report.append(f"Dette probable: ${stats['dette_probable']:,.2f}")
    report.append(f"DETTE TOTALE ESTIMEE: ${stats['dette_totale_estimee']:,.2f}")
    report.append("")

    # Informations client
    report.append("-" * 70)
    report.append("  INFORMATIONS CLIENT")
    report.append("-" * 70)
    report.append("")
    report.append(f"Nom: {client_info.get('nom', 'Non disponible')}")
    report.append(f"Email: {client_info.get('email', 'Non disponible')}")
    report.append(f"Telephone: {client_info.get('telephone', 'Non disponible')}")
    report.append(f"Institution: {client_info.get('institution', 'Non disponible')}")
    report.append(f"Compte: {client_info.get('compte', 'Non disponible')}")
    report.append(f"Solde: ${client_info.get('solde', 0):,.2f}")
    report.append("")

    # Capacite financiere
    report.append("-" * 70)
    report.append("  CAPACITE FINANCIERE")
    report.append("-" * 70)
    report.append("")
    report.append(f"Revenu mensuel estime: ${monthly_income:,.2f}")
    report.append(f"Capacite mensuelle (50%): ${capacite_mensuelle:,.2f}")
    report.append(f"Paiement max/semaine: ${capacite_hebdo:,.2f}")
    report.append("")

    # Preteurs confirmes (Liste OPC)
    if preteurs_confirmes:
        report.append("-" * 70)
        report.append("  PRETEURS CONFIRMES [LISTE OFFICIELLE OPC]")
        report.append("  Confiance: TRES ELEVEE (95-100%)")
        report.append("-" * 70)
        report.append("")

        for preteur in sorted(preteurs_confirmes, key=lambda x: x['total_paye'], reverse=True):
            report.append(f"  {preteur['nom']}")
            report.append(f"  Source: {preteur['source']}")
            report.append(f"  Score: {preteur['score']}/100 | Confiance: {preteur['confidence']}")
            report.append(f"  Total paye: ${preteur['total_paye']:,.2f}")
            report.append(f"  Transactions: {len(preteur['transactions'])}")

            # Afficher les 3 dernieres transactions
            for trans in preteur['transactions'][-3:]:
                report.append(f"    - {trans['date']}: ${trans['amount']:,.2f}")
            report.append("")

    # Preteurs probables (Regles)
    if preteurs_probables:
        report.append("-" * 70)
        report.append("  PRETEURS PROBABLES [REGLES INTELLIGENTES]")
        report.append("  Confiance: ELEVEE (60-79%)")
        report.append("-" * 70)
        report.append("")

        for preteur in sorted(preteurs_probables, key=lambda x: x['score'], reverse=True):
            report.append(f"  {preteur['nom'][:50]}")
            report.append(f"  Source: REGLES INTELLIGENTES")
            report.append(f"  Score: {preteur['score']}/100 | Confiance: {preteur['confidence']}")
            report.append(f"  Total paye: ${preteur['total_paye']:,.2f}")
            report.append(f"  Raisons:")
            for reason in preteur['reasons'][:3]:
                report.append(f"    - {reason}")
            report.append("")

    # Preteurs possibles (A verifier)
    if preteurs_possibles:
        report.append("-" * 70)
        report.append("  PRETEURS POSSIBLES [A VERIFIER]")
        report.append("  Confiance: MOYENNE (40-59%)")
        report.append("-" * 70)
        report.append("")

        for preteur in sorted(preteurs_possibles, key=lambda x: x['score'], reverse=True):
            report.append(f"  {preteur['nom'][:50]}")
            report.append(f"  Score: {preteur['score']}/100 | Action: {preteur['action']}")
            report.append(f"  Total: ${preteur['total_paye']:,.2f} ({len(preteur['transactions'])} transactions)")
            report.append("")

    # Exclusions
    if exclusions:
        report.append("-" * 70)
        report.append("  EXCLUSIONS AUTOMATIQUES [FAUX POSITIFS ELIMINES]")
        report.append("-" * 70)
        report.append("")

        for excl in exclusions:
            report.append(f"  {excl['nom']}")
            report.append(f"  Raison: {excl['reasons'][0] if excl['reasons'] else 'Exclusion'}")
            report.append(f"  Montant exclu: ${excl['total_paye']:,.2f}")
            report.append("")

    # ==========================================================================
    # NOUVEAU SYSTÈME DE SCORING AVEC KNOCK-OUTS
    # ==========================================================================

    # Extraire NSF et overdrafts des transactions
    nsf_30j, overdraft_90j = extraire_nsf_et_overdraft(transactions, inverite_data)

    # Calculer les variables pour l'évaluation
    total_dette = stats['dette_totale_estimee']
    nb_preteurs = stats['nb_preteurs_confirmes'] + stats['nb_preteurs_probables']
    ratio_dette_revenu = total_dette / monthly_income if monthly_income > 0 else 1.0
    budget_disponible = capacite_mensuelle - total_dette

    # Préparer les données pour l'évaluation
    data_evaluation = {
        'revenus_mensuels': monthly_income,
        'total_paiements_prets': total_dette,
        'budget_disponible': budget_disponible,
        'nb_preteurs_confirmes': nb_preteurs,
        'nsf_30_jours': nsf_30j,
        'overdraft_90_jours': overdraft_90j
    }

    # Vérifier les knock-outs
    knock_outs_result = verifier_knock_outs(data_evaluation)

    # Calculer le score
    score_result = calculer_score_avance(data_evaluation, knock_outs_result)

    score_global = score_result['score']
    niveau_risque = score_result['niveau_risque']

    # Section KNOCK-OUTS (si détectés)
    if knock_outs_result['has_knock_outs']:
        report.append("")
        report.append("=" * 70)
        report.append("  ⛔ KNOCK-OUTS DÉTECTÉS - REFUS AUTOMATIQUE")
        report.append("=" * 70)
        report.append("")

        for ko in knock_outs_result['knock_outs']:
            report.append(f"  🔴 {ko['message']}")

        report.append("")
        report.append(f"  Nombre total de knock-outs: {knock_outs_result['nb_knock_outs']}")
        report.append("  DÉCISION: REFUS - Client ne peut pas payer mathématiquement")
        report.append("")

    # Section SCORE GLOBAL
    report.append("")
    report.append("=" * 70)
    report.append(f"  SCORE GLOBAL: {score_global}/100 - {niveau_risque}")
    report.append("=" * 70)
    report.append("")

    if score_result.get('penalites'):
        report.append("  Pénalités appliquées:")
        for penalite in score_result['penalites']:
            report.append(f"    • {penalite}")
        report.append("")

    # Offres disponibles
    report.append("-" * 70)
    report.append("  OFFRES DISPONIBLES")
    report.append("-" * 70)
    report.append("")

    capacite_50_pct = capacite_mensuelle  # Capacité max = 50% du revenu mensuel

    for montant in [500, 1000, 1500, 2000, 2500, 3000]:
        if montant in paiements:
            paiement_sem = paiements[montant]["sem"]

            # NOUVELLE LOGIQUE: Probabilité réaliste
            chance = calculer_probabilite_realiste(
                score_global=score_global,
                ratio_dette=ratio_dette_revenu,
                nb_preteurs=nb_preteurs,
                montant_demande=montant,
                capacite_max=capacite_50_pct
            )

            report.append(f"  ${montant:,} - Paiement: ${paiement_sem:.2f}/sem - Chance: {chance:.1f}%")

    report.append("")
    report.append(f"  [Score client: {score_global}/100 | Ratio dette: {ratio_dette_revenu*100:.1f}% | Prêteurs: {nb_preteurs}]")
    report.append("")

    # Recommandations
    report.append("-" * 70)
    report.append("  RECOMMANDATIONS")
    report.append("-" * 70)
    report.append("")

    total_dette = stats['dette_totale_estimee']
    nb_preteurs = stats['nb_preteurs_confirmes'] + stats['nb_preteurs_probables']

    if nb_preteurs == 0:
        report.append("  Aucun preteur actif detecte - Client eligible")
    elif nb_preteurs <= 2 and total_dette < 2000:
        report.append("  Risque FAIBLE - Client eligible avec prudence")
    elif nb_preteurs <= 4 and total_dette < 5000:
        report.append("  Risque MOYEN - Verifier capacite de remboursement")
    else:
        report.append("  Risque ELEVE - Evaluation approfondie requise")

    report.append("")

    # Pied de page
    report.append("=" * 70)
    report.append("  FIN DU RAPPORT")
    report.append("=" * 70)
    report.append("")
    report.append("Systeme: SAR Analysis Hybride v2.0")
    report.append("Listes: OPC Quebec (414 preteurs) + Regles intelligentes")
    report.append(f"Genere le: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    report.append("")

    return "\n".join(report)


# =============================================================================
# FONCTIONS AUXILIAIRES (importees de analyser.py)
# =============================================================================

def load_paiements_excel():
    """Charge la grille de paiements depuis le fichier Excel"""
    paiements = {}
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:
                montant = int(row[0])
                paiements[montant] = {
                    "sem": float(row[1]),
                    "2sem": float(row[2])
                }
    except Exception as e:
        print(f"Erreur lecture Excel: {e}")
        paiements = {
            500: {"sem": 47.77, "2sem": 95.52},
            1000: {"sem": 75.13, "2sem": 150.25},
            1500: {"sem": 96.18, "2sem": 192.37},
            2000: {"sem": 117.23, "2sem": 234.47},
            2500: {"sem": 138.29, "2sem": 276.57},
            3000: {"sem": 159.34, "2sem": 318.68},
        }
    return paiements


def calculate_monthly_income_from_transactions(transactions):
    """Calcule le revenu mensuel depuis les transactions"""
    if not transactions:
        return 0.0

    monthly_income = defaultdict(float)

    for trans in transactions:
        amount = trans.get('amount', trans.get('Montant', 0))
        date = trans.get('date', trans.get('Date'))
        description = trans.get('description', trans.get('Description', '')).upper()

        if amount > 0 and date:  # Credits seulement
            if isinstance(date, str):
                try:
                    date = datetime.strptime(date[:10], "%Y-%m-%d")
                except:
                    continue

            month_key = date.strftime("%Y-%m")

            # Verifier si c'est un depot de paie
            if any(kw in description for kw in PAYROLL_KEYWORDS):
                monthly_income[month_key] += amount

    if not monthly_income:
        # Compter tous les credits si pas de paie detectee
        for trans in transactions:
            amount = trans.get('amount', trans.get('Montant', 0))
            date = trans.get('date', trans.get('Date'))

            if amount > 0 and date:
                if isinstance(date, str):
                    try:
                        date = datetime.strptime(date[:10], "%Y-%m-%d")
                    except:
                        continue

                month_key = date.strftime("%Y-%m")
                monthly_income[month_key] += amount

    if not monthly_income:
        return 0.0

    return sum(monthly_income.values()) / len(monthly_income)


def extract_fulltext_from_json(data):
    """Extrait le texte brut fullText/rawText depuis le JSON"""
    if isinstance(data, dict):
        if "content" in data and isinstance(data["content"], dict):
            if "fullText" in data["content"]:
                return data["content"]["fullText"]
            if "rawText" in data["content"]:
                return data["content"]["rawText"]
        if "fullText" in data:
            return data["fullText"]
        if "rawText" in data:
            return data["rawText"]
    return None


def extract_tables_from_json(data):
    """Extrait les tables depuis le JSON"""
    if isinstance(data, dict):
        if "content" in data and isinstance(data["content"], dict):
            if "tables" in data["content"]:
                return data["content"]["tables"]
        if "tables" in data:
            return data["tables"]
    return None


def parse_transactions_from_tables(tables):
    """Parse les transactions depuis les tables JSON"""
    transactions = []

    if not tables:
        return transactions

    for table in tables:
        headers = [h.lower() if h else "" for h in table.get("headers", [])]
        rows = table.get("rows", [])

        date_idx = -1
        details_idx = -1
        credit_idx = -1
        debit_idx = -1

        for i, h in enumerate(headers):
            if h == "date":
                date_idx = i
            elif h == "details":
                details_idx = i
            elif h == "credit":
                credit_idx = i
            elif h == "debit":
                debit_idx = i

        if date_idx == -1:
            continue

        for row in rows:
            if len(row) < 3:
                continue

            date_str = str(row[date_idx]).strip() if date_idx < len(row) else ""
            if not date_str or not re.match(r'\d{4}-\d{2}-\d{2}', date_str):
                continue

            try:
                trans_date = datetime.strptime(date_str, "%Y-%m-%d")
            except:
                continue

            description = str(row[details_idx]).strip().upper() if details_idx >= 0 and details_idx < len(row) else ""

            credit = 0.0
            debit = 0.0

            if credit_idx >= 0 and credit_idx < len(row):
                credit_str = str(row[credit_idx]).replace("$", "").replace(",", "").strip()
                try:
                    credit = float(credit_str) if credit_str else 0.0
                except:
                    credit = 0.0

            if debit_idx >= 0 and debit_idx < len(row):
                debit_str = str(row[debit_idx]).replace("$", "").replace(",", "").strip()
                try:
                    debit = float(debit_str) if debit_str else 0.0
                except:
                    debit = 0.0

            if credit > 0:
                amount = credit
                trans_type = "credit"
            elif debit > 0:
                amount = -debit
                trans_type = "debit"
            else:
                continue

            transactions.append({
                "date": trans_date,
                "description": description,
                "amount": amount,
                "type": trans_type
            })

    return transactions


def parse_client_info_from_tables(tables):
    """Parse les informations client depuis les tables"""
    info = {
        "nom": "",
        "adresse": "",
        "email": "",
        "telephone": "",
        "institution": "",
        "compte": "",
        "solde": 0.0,
        "revenu_mensuel": 0.0
    }

    if not tables:
        return info

    for table in tables:
        rows = table.get("rows", [])
        for row in rows:
            if len(row) >= 2:
                key = str(row[0]).strip().lower()
                value = str(row[1]).strip()

                if key == "name" and value and not info["nom"]:
                    info["nom"] = value
                elif key == "email" and value and not info["email"]:
                    info["email"] = value
                elif key == "phone" and value and not info["telephone"]:
                    info["telephone"] = value

    return info


# =============================================================================
# TESTS AUTOMATIQUES
# =============================================================================

def run_tests():
    """Execute les tests automatiques du systeme hybride"""

    print("=" * 70)
    print("  TESTS AUTOMATIQUES - SYSTEME HYBRIDE V2.0")
    print("=" * 70)
    print()

    test_cases = [
        # Format: (description, amount, expected_min_score, expected_source, test_name)

        # === TESTS LISTE OFFICIELLE OPC ===
        ("MONEY MART MONTREAL", 300, 90, "LISTE_OPC", "Preteur OPC - Money Mart"),
        ("ADVANCE CREDIT TM", 500, 90, "LISTE_OPC", "Preteur OPC - Advance Credit"),
        ("FLEXITI FINANCIAL", 1200, 90, "LISTE_OPC", "Preteur OPC - Flexiti"),
        ("KREDIT PRET INC", 800, 90, "LISTE_OPC", "Preteur OPC - Kredit Pret"),
        ("LENDORA SERVICES", 600, 90, "LISTE_OPC", "Preteur OPC - Lendora"),
        ("FAIRSTONE FINANCIAL", 450, 90, "LISTE_OPC", "Preteur OPC - Fairstone"),

        # === TESTS LISTE COMPLEMENTAIRE ===
        ("CASH MONEY", 200, 85, "LISTE_COMPLEMENTAIRE", "Preteur Complementaire - Cash Money"),

        # === TESTS EXCLUSIONS ===
        ("BENEVA ASSURANCE", 851, 0, "EXCLUSION", "Exclusion - Assurance Beneva"),
        ("GIGADAT INC", 30497, 0, "EXCLUSION", "Exclusion - Casino Gigadat"),
        ("TIM HORTONS #6364", 15, 0, "EXCLUSION", "Exclusion - Commerce Tim Hortons"),
        ("RAYMOND CHABOT SYNDIC", 500, 0, "EXCLUSION", "Exclusion - Syndic"),

        # === TESTS REGLES INTELLIGENTES ===
        ("PMT CREDIT SERVICE 123", 500, 40, "REGLES", "Regles - PMT + CREDIT + montant rond"),
        ("VIR LOAN COMPANY ABC", 300, 40, "REGLES", "Regles - VIR + LOAN"),
        ("FINANCE EXPRESS INC", 1000, 40, "REGLES", "Regles - FINANCE + EXPRESS"),

        # === TESTS NON-PRETEURS ===
        ("SALAIRE EMPLOYEUR ABC", 3000, 0, "AUCUNE", "Non-preteur - Salaire"),
        ("VIREMENT PERSONNEL JEAN", 100, 0, "AUCUNE", "Non-preteur - Virement personnel"),
        ("DEPOT PAIE COMPAGNIE XYZ", 2500, 0, "AUCUNE", "Non-preteur - Depot paie"),
        ("REMBOURSEMENT IMPOT 2024", 1500, 0, "AUCUNE", "Non-preteur - Remboursement impot"),

        # === TESTS VIR INTERAC (PAS des preteurs!) ===
        ("VIR INTERAC ENVOYE JEFFFFFF 202532404482", 300, 0, "AUCUNE", "VIR INTERAC - Transfert personnel"),
        ("VIR INTERAC ENVOYE LUC PROPRIETAIRE 2025", 1000, 0, "AUCUNE", "VIR INTERAC - Loyer"),
        ("VIR INTERAC ENVOYE COIFFEUR JOHANNE 2025", 60, 0, "AUCUNE", "VIR INTERAC - Service coiffeur"),
        ("VIR INTERAC EFFECTUE PAYPER INC 2025", 100, 0, "AUCUNE", "VIR INTERAC - Paiement service"),
        ("VIREMENT INTERAC MARIE 2025", 200, 0, "AUCUNE", "VIREMENT INTERAC - Transfert"),
    ]

    passed = 0
    failed = 0

    for desc, amount, expected_min_score, expected_source, test_name in test_cases:
        result = calculate_lender_score_hybrid(desc, amount, [])

        # Verifier le score
        if expected_source in ['EXCLUSION', 'AUCUNE']:
            score_ok = result['score'] <= expected_min_score
        else:
            score_ok = result['score'] >= expected_min_score

        # Verifier la source
        source_ok = result['source'] == expected_source

        # Pour les regles, accepter aussi si score est suffisant
        if expected_source == 'REGLES' and result['score'] >= expected_min_score:
            source_ok = result['source'] in ['REGLES', 'LISTE_OPC', 'LISTE_COMPLEMENTAIRE']

        if score_ok and source_ok:
            print(f"  [PASS] {test_name}")
            print(f"         Score: {result['score']}, Source: {result['source']}")
            passed += 1
        else:
            print(f"  [FAIL] {test_name}")
            print(f"         Attendu: score {'<=' if expected_source in ['EXCLUSION', 'AUCUNE'] else '>='} {expected_min_score}, source = {expected_source}")
            print(f"         Obtenu: score = {result['score']}, source = {result['source']}")
            if result['reasons']:
                print(f"         Raisons: {', '.join(result['reasons'][:2])}")
            failed += 1
        print()

    print("=" * 70)
    print(f"  RESULTATS: {passed} PASS, {failed} FAIL")
    print(f"  Taux de reussite: {passed/(passed+failed)*100:.1f}%")
    print("=" * 70)

    return failed == 0


# =============================================================================
# TRAITEMENT FICHIER JSON
# =============================================================================

def process_json_file_hybrid(filepath, paiements):
    """Traite un fichier JSON avec le systeme hybride"""
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            data = json.load(f)

        filename = os.path.basename(filepath)

        # Extraire donnees
        tables = extract_tables_from_json(data)
        fulltext = extract_fulltext_from_json(data)

        # Parser client info - supporter plusieurs formats
        client_info = {
            "nom": "", "adresse": "", "email": "", "telephone": "",
            "institution": "", "compte": "", "solde": 0.0, "revenu_mensuel": 0.0
        }

        if tables:
            client_info = parse_client_info_from_tables(tables)
        elif "customerInfo" in data:
            # Format tempData de route.ts (Inverite converti)
            ci = data["customerInfo"]
            client_info["nom"] = ci.get("name", "") or ci.get("referenceid", "")
        elif "name" in data or "referenceid" in data:
            # Format Inverite direct - utiliser name ou referenceid comme fallback
            client_info["nom"] = (data.get("name", "") or "").strip() or data.get("referenceid", "")

        if "extraction" in data:
            client_info["institution"] = data["extraction"].get("institution", "")
        if "Institution" in data:
            client_info["institution"] = data["Institution"].get("Title", "")

        # Format Inverite: Extraire infos depuis accounts[]
        if "accounts" in data and len(data.get("accounts", [])) > 0:
            acc = data["accounts"][0]
            # Solde (current_balance ou available_balance)
            if not client_info.get("solde"):
                solde = acc.get("current_balance") or acc.get("available_balance") or 0
                client_info["solde"] = float(solde or 0)
            # Institution (description du compte)
            if not client_info.get("institution"):
                client_info["institution"] = acc.get("account_description", "")
            # Numéro de compte
            if not client_info.get("compte"):
                client_info["compte"] = acc.get("account", "")
            # Revenu mensuel depuis statistics
            stats = acc.get("statistics", {})
            revenu_employer = float(stats.get("average_monthly_employer_income", 0) or 0)
            revenu_govt = float(stats.get("average_monthly_govt_income", 0) or 0)
            revenu_total = revenu_employer + revenu_govt
            if revenu_total > 0:
                client_info["revenu_mensuel"] = revenu_total

        # Parser transactions - supporter plusieurs formats
        transactions = []

        # Format 1: Tables standard
        if tables:
            transactions = parse_transactions_from_tables(tables)

        # Format 2: Array 'transactions' direct (tempData de route.ts)
        if not transactions and "transactions" in data:
            for t in data["transactions"]:
                date_str = t.get("Date", t.get("date", ""))
                desc = t.get("Description", t.get("description", ""))
                amount = t.get("Amount", t.get("amount", t.get("Montant", 0)))
                trans_type = t.get("Type", t.get("type", ""))
                category = t.get("category", "")

                # Convertir amount en float si string
                if isinstance(amount, str):
                    amount = float(amount.replace(",", "").replace("$", "")) if amount else 0

                transactions.append({
                    "date": date_str,
                    "description": desc,
                    "amount": amount,
                    "type": trans_type,
                    "category": category
                })

        # Format 3: Inverite direct avec accounts[].transactions
        if not transactions and "accounts" in data:
            for acc in data.get("accounts", []):
                for t in acc.get("transactions", []):
                    date_str = t.get("date", "")
                    # Inverite utilise 'details' pas 'description'
                    desc = t.get("details", t.get("description", ""))
                    credit = float(t.get("credit", "0") or "0")
                    debit = float(t.get("debit", "0") or "0")
                    category = t.get("category", "")

                    if credit > 0:
                        amount = credit
                        trans_type = "credit"
                    else:
                        amount = -debit
                        trans_type = "debit"

                    transactions.append({
                        "date": date_str,
                        "description": desc,
                        "amount": amount,
                        "type": trans_type,
                        "category": category
                    })

        if not transactions:
            return f"Aucune transaction trouvee dans {filepath}", None

        # Detection hybride
        detection_result = detect_lenders_hybrid(transactions)

        # Generer rapport
        report = generate_report_enrichi(client_info, transactions, detection_result, paiements, filename, data)

        return report, client_info.get("nom", filename)

    except json.JSONDecodeError as e:
        return f"Erreur JSON dans {filepath}: {e}", None
    except Exception as e:
        return f"Erreur traitement {filepath}: {e}", None


# =============================================================================
# MAIN
# =============================================================================

def main():
    """Fonction principale"""

    print()
    print("=" * 70)
    print("  SAR ANALYSIS - SYSTEME HYBRIDE V2.0")
    print("  Liste OPC (414 preteurs) + Regles intelligentes")
    print("=" * 70)
    print()

    # Arguments
    if len(sys.argv) > 1 and sys.argv[1] == "--test":
        # Mode test
        success = run_tests()
        sys.exit(0 if success else 1)

    # Mode fichier unique (appel depuis API)
    if len(sys.argv) > 1 and sys.argv[1] != "--test" and os.path.isfile(sys.argv[1]):
        filepath = sys.argv[1]
        print(f"Mode API: Traitement de {filepath}")

        # Charger grille paiements
        paiements = load_paiements_excel()

        # Verifier dossier output
        if not os.path.exists(RAPPORTS_OUTPUT_DIR):
            os.makedirs(RAPPORTS_OUTPUT_DIR)

        # Traiter le fichier unique
        report, client_name = process_json_file_hybrid(filepath, paiements)

        if client_name:
            safe_name = re.sub(r'[^\w\s-]', '', client_name or "api")
            safe_name = re.sub(r'[-\s]+', '_', safe_name).strip('_')
            if not safe_name:
                safe_name = "api_request"

            output_filename = f"rapport_hybride_{safe_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
            output_path = os.path.join(RAPPORTS_OUTPUT_DIR, output_filename)

            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(report)

            print(f"Rapport genere: {output_filename}")
        else:
            print(f"Erreur: {report}")

        return

    # Verifier les dossiers
    if not os.path.exists(JSON_INPUT_DIR):
        os.makedirs(JSON_INPUT_DIR)
        print(f"Dossier cree: {JSON_INPUT_DIR}")

    if not os.path.exists(RAPPORTS_OUTPUT_DIR):
        os.makedirs(RAPPORTS_OUTPUT_DIR)
        print(f"Dossier cree: {RAPPORTS_OUTPUT_DIR}")

    # Charger grille paiements
    paiements = load_paiements_excel()
    print(f"Grille de paiements chargee: {len(paiements)} montants")

    # Afficher stats listes
    print()
    print(f"Listes chargees:")
    print(f"  - Preteurs OPC: {len(PRETEURS_OFFICIELS_OPC)}")
    print(f"  - Preteurs complementaires: {len(PRETEURS_COMPLEMENTAIRES)}")
    print(f"  - Assurances (exclusion): {len(ASSURANCES_TOUTES)}")
    print(f"  - Syndics (exclusion): {len(SYNDICS_TOUS)}")
    print(f"  - Casinos (exclusion): {len(CASINOS_JEUX)}")
    print(f"  - Commerces (exclusion): {len(COMMERCES_SERVICES)}")
    print()

    # Trouver fichiers JSON
    json_files = glob.glob(os.path.join(JSON_INPUT_DIR, "*.json"))

    if not json_files:
        print("ATTENTION: Aucun fichier JSON trouve dans json-input/")
        print(f"Placez vos fichiers JSON dans: {JSON_INPUT_DIR}")
        print()
        print("Lancement des tests automatiques...")
        print()
        run_tests()
        return

    print(f"Fichiers JSON trouves: {len(json_files)}")
    print("-" * 70)

    # Traiter chaque fichier
    reports_generated = 0

    for filepath in json_files:
        filename = os.path.basename(filepath)
        print(f"Traitement: {filename}...")

        report, client_name = process_json_file_hybrid(filepath, paiements)

        if client_name:
            safe_name = re.sub(r'[^\w\s-]', '', client_name or filename)
            safe_name = re.sub(r'[-\s]+', '_', safe_name).strip('_')
            if not safe_name:
                safe_name = filename.replace('.json', '')

            output_filename = f"rapport_hybride_{safe_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
            output_path = os.path.join(RAPPORTS_OUTPUT_DIR, output_filename)

            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(report)

            print(f"  -> Rapport genere: {output_filename}")
            reports_generated += 1
        else:
            print(f"  -> Erreur: {report}")

    print()
    print("-" * 70)
    print(f"TERMINE! {reports_generated} rapport(s) genere(s)")
    print(f"Rapports disponibles dans: {RAPPORTS_OUTPUT_DIR}")
    print()


if __name__ == "__main__":
    main()
